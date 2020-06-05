# coding:utf8
import os
import sys

import openpyxl

base_path = os.getcwd()
sys.path.append(base_path)


class HandExcel:

    # 加载excel
    def load_excel(self):
        open_excel = openpyxl.load_workbook("../Case/ApiAuto.xlsx")  # 首先对这个对象进行获取load本工程下得excel数据
        return open_excel

    # 加载所有sheet得内容
    def get_sheet_data(self, index=None):
        sheet_name = self.load_excel().sheetnames
        if index is None:
            index = 0
        data = self.load_excel()[sheet_name[index]]
        return data

    # 获取一个单元格得内容是由’行‘和‘列’决定
    def get_cell_value(self, row, cols):
        # get_sheet_data()使用默认值0
        data = self.get_sheet_data().cell(row=row, column=cols).value
        return data

    # 获取sheet页中得case行数
    def get_rows(self):
        row = self.get_sheet_data().max_row
        return row

    # 某一行得内容得list
    def get_rows_value(self, row):
        row_list = []
        for i in self.get_sheet_data()[row]:
            row_list.append(i.value)
        return row_list

    def excel_write_data(self, row, cols, value):
        """
        :param row:
        :param cols:
        :param value:
        :return: 写入数据
        """
        workbook = self.load_excel()
        write = workbook.active  # 启动，激活
        write.cell(row, cols, value)
        workbook.save("../Case/ApiAuto.xlsx")

    def get_columns_value(self, key=None):
        """
        :param key:
        :return: 获取某一列得数据
        """

        columns_list = []
        if key is None:
            key = 'A'
        columns_list_data = self.get_sheet_data()[key]
        for i in columns_list_data:
            columns_list.append(i.value)
        return columns_list

    def get_rows_number(self, case_id):
        """
        :param case_id:
        :return: 根据case_id获取某一行得数据
        """
        num = 1  # 定义行号
        cols_data = self.get_columns_value()
        for col_data in cols_data:
            if case_id == col_data:
                return num
            num = num + 1
        return num

    def get_excel_data(self):
        """
        :return: 获取excel里面所有的数据
        """

        data_list = []
        for i in range(self.get_rows()):
            data_list.append(self.get_rows_value(i + 2))

        return data_list


excel_data = HandExcel()

if __name__ == "__main__":
    handle = HandExcel()
    print(handle.get_cell_value(2, 5))
    print(handle.get_rows_value(2))  # 打印第2行内所有内容
    print(handle.get_rows_number("test_004"))
